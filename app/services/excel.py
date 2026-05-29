from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.schemas.orders import ReportOut


HEADER_FILL = PatternFill("solid", fgColor="C7B7F3")
HEADER_FONT = Font(color="2B0B5F", bold=True)
DISPLAY_NAMES = {
    "Cafe": "Café",
    "Con jamon y queso": "Con jamón y queso",
    "Solo jamon": "Solo jamón",
    "Dietetico": "Dietético",
    "Dietetico adicional": "Dietético adicional",
    "Pina": "Piña",
    "Yogurt pequeno": "Yogurt pequeño",
}


def display_name(value: str | None) -> str | None:
    return DISPLAY_NAMES.get(value or "", value)


def style_header(sheet, row: int) -> None:
    for cell in sheet[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def write_counter_section(sheet, row: int, title: str, data: dict[str, int]) -> int:
    sheet.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 1
    sheet.cell(row=row, column=1, value="Concepto")
    sheet.cell(row=row, column=2, value="Cantidad")
    style_header(sheet, row)
    row += 1
    if data:
        for label, value in data.items():
            sheet.cell(row=row, column=1, value=display_name(label))
            sheet.cell(row=row, column=2, value=value)
            row += 1
    else:
        sheet.cell(row=row, column=1, value="Sin datos")
        sheet.cell(row=row, column=2, value=0)
        row += 1
    return row + 1


def write_extra_details(sheet, row: int, details: list[dict]) -> int:
    sheet.cell(row=row, column=1, value="Adicionales por huésped").font = Font(bold=True)
    row += 1
    headers = ["Huésped", "DNI", "Adicional", "Cantidad", "Precio unitario", "Total"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=column, value=header)
    style_header(sheet, row)
    row += 1
    if details:
        for detail in details:
            sheet.cell(row=row, column=1, value=detail.get("guest_name"))
            sheet.cell(row=row, column=2, value=detail.get("document"))
            sheet.cell(row=row, column=3, value=display_name(detail.get("extra_name")))
            sheet.cell(row=row, column=4, value=detail.get("quantity"))
            sheet.cell(row=row, column=5, value=detail.get("unit_price"))
            sheet.cell(row=row, column=6, value=detail.get("total"))
            sheet.cell(row=row, column=5).number_format = '"S/" #,##0.00'
            sheet.cell(row=row, column=6).number_format = '"S/" #,##0.00'
            row += 1
    else:
        sheet.cell(row=row, column=1, value="Sin datos")
        row += 1
    return row + 1


def report_to_excel(report: ReportOut) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte diario"
    sheet["A1"] = f"QR System - Reporte {report.date}"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A3"] = "Total pedidos"
    sheet["B3"] = report.total_orders

    row = 5
    row = write_counter_section(sheet, row, "Atendidos por origen", report.attended_by_origin)
    row = write_counter_section(sheet, row, "Tipos de desayuno", report.breakfast_types)
    row = write_extra_details(sheet, row, report.extra_details)
    row = write_counter_section(sheet, row, "Adicionales más pedidos", report.extras)
    row = write_counter_section(sheet, row, "Horas pico", report.peak_hours)
    write_counter_section(sheet, row, "Motivos de cancelación", report.cancellation_reasons)

    widths = {
        "A": 32,
        "B": 18,
        "C": 30,
        "D": 14,
        "E": 18,
        "F": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
